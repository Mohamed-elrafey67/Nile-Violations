import json
from django.db import models
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Sum, Count, Q
from django.utils import timezone
from .models import Violation, Governorate, District, UserProfile, ViolationImage, ViolationNote

_GEO_ALL = None
_GOVS    = None

def _load_geo():
    global _GEO_ALL, _GOVS
    if _GEO_ALL is None:
        with open(settings.GEO_JSON_PATH, 'r', encoding='utf-8') as f:
            _GEO_ALL = json.load(f)
    if _GOVS is None:
        with open(settings.GOVS_JSON_PATH, 'r', encoding='utf-8') as f:
            _GOVS = json.load(f)

def get_role(user):
    if not user.is_authenticated: return None
    if user.is_superuser: return 'manager'
    try: return user.profile.role
    except: return 'viewer'

# ── AUTH ──────────────────────────────────────────────────────────
def login_view(request):
    error = None
    if request.method == 'POST':
        username = request.POST.get('username','')
        u = authenticate(request, username=username,
                         password=request.POST.get('password',''))
        if u:
            login(request, u)
            ip = (request.META.get('HTTP_X_FORWARDED_FOR','').split(',')[0].strip()
                  or request.META.get('REMOTE_ADDR',''))
            from .models import AuditLog
            AuditLog.objects.create(user=u, action='login',
                target='تسجيل دخول ناجح', ip_address=ip or None)
            return redirect('map')
        error = 'اسم المستخدم أو كلمة المرور غير صحيحة'
        ip = (request.META.get('HTTP_X_FORWARDED_FOR','').split(',')[0].strip()
              or request.META.get('REMOTE_ADDR',''))
        from .models import AuditLog
        AuditLog.objects.create(user=None, action='login_fail',
            target=f'محاولة فاشلة: {username}', ip_address=ip or None)
    return render(request, 'violations/login.html', {'error': error})

def logout_view(request):
    if request.user.is_authenticated:
        log_action(request, 'logout', 'تسجيل خروج')
    logout(request)
    return redirect('login')

# ── MAIN MAP ──────────────────────────────────────────────────────
@login_required(login_url='login')
def map_view(request):
    role = get_role(request.user)
    govs_with_data = Governorate.objects.filter(has_data=True).order_by('name_ar')
    overall = Violation.objects.filter(status='approved').aggregate(
        total_count=Count('id'), total_area=Sum('area_total'),
        geo_count=Count('id', filter=Q(geo_exact=True)),
    )
    pending_count = Violation.objects.filter(status='pending').count() if role in ('supervisor','manager') else 0

    # كل محافظات مصر من ملف الجيو (بغض النظر عن وجود بيانات)
    _load_geo()
    all_egypt_govs = sorted(_GOVS, key=lambda g: g['name_ar'])

    context = {
        'govs_with_data': govs_with_data, 'overall': overall,
        'role': role, 'pending_count': pending_count,
        'all_govs': all_egypt_govs,
        'user': request.user,
    }
    return render(request, 'violations/map.html', context)

# ── GEO API ───────────────────────────────────────────────────────
@login_required(login_url='login')
def geo_gov_api(request, gov_pcode):
    _load_geo()
    polys = [v for v in _GEO_ALL if v['gov_pcode'] == gov_pcode]
    counts = dict(Violation.objects.filter(governorate__pcode=gov_pcode, status='approved')
                  .values_list('village_pcode').annotate(c=Count('id')))
    areas  = dict(Violation.objects.filter(governorate__pcode=gov_pcode, status='approved')
                  .values_list('village_pcode').annotate(s=Sum('area_total')))
    for p in polys:
        p['violation_count'] = counts.get(p['pcode'], 0)
        p['total_area']      = round(areas.get(p['pcode'], 0) or 0, 1)
    return JsonResponse({'polygons': polys, 'gov_pcode': gov_pcode})

# ── VIOLATIONS API ────────────────────────────────────────────────
@login_required(login_url='login')
def violations_api(request):
    role = get_role(request.user)
    qs = Violation.objects.select_related('governorate').all()
    if role not in ('supervisor', 'manager'):
        qs = qs.filter(status='approved')

    gov=request.GET.get('gov',''); district=request.GET.get('district','')
    village=request.GET.get('village',''); desc=request.GET.get('description','')
    min_area=request.GET.get('min_area',''); search=request.GET.get('search','')
    pcode=request.GET.get('pcode',''); status=request.GET.get('status','')

    if gov:      qs = qs.filter(governorate__pcode=gov)
    if district: qs = qs.filter(district_name=district)
    if village:  qs = qs.filter(village=village)
    if desc:     qs = qs.filter(description__icontains=desc)
    if pcode:    qs = qs.filter(village_pcode=pcode)
    if status and role in ('supervisor','manager'): qs = qs.filter(status=status)
    if min_area:
        try: qs = qs.filter(area_total__gte=float(min_area))
        except: pass
    if search:
        qs = qs.filter(Q(occupant__icontains=search)|Q(code__icontains=search)|
                       Q(village__icontains=search)|Q(basin__icontains=search))

    summary = qs.aggregate(count=Count('id'), total_area=Sum('area_total'))
    records = list(qs.values('id','code','governorate__name_ar','governorate__pcode',
        'district_name','village','village_pcode','occupant','basin','description',
        'area_outside','area_public','area_nile_bank','area_total',
        'latitude','longitude','geo_exact','status','field_notes')[:500])
    return JsonResponse({'records': records, 'summary': summary, 'role': role})

# ── FILTER OPTIONS ────────────────────────────────────────────────
@login_required(login_url='login')
def filter_options_api(request):
    gov      = request.GET.get('gov', '')
    district = request.GET.get('district', '')

    # Districts: from geo JSON (all villages) + DB (existing violations)
    _load_geo()
    geo_districts = []
    geo_villages  = []

    if gov and _GEO_ALL:
        # Get districts from geo data
        dist_set = set()
        for v in _GEO_ALL:
            if v['gov_pcode'] == gov:
                dist_set.add(v['district_ar'])
        geo_districts = sorted(dist_set)

        # Get villages filtered by district
        if district:
            for v in _GEO_ALL:
                if v['gov_pcode'] == gov and v['district_ar'] == district:
                    geo_villages.append(v['name_ar'])
            geo_villages = sorted(geo_villages)

    # Also get existing violation descriptions
    qs = Violation.objects.all()
    if gov:      qs = qs.filter(governorate__pcode=gov)
    if district: qs = qs.filter(district_name=district)
    descs = list(qs.values_list('description', flat=True).distinct().order_by('description'))

    # Merge DB districts with geo districts
    db_districts = list(qs.values_list('district_name', flat=True).distinct())
    all_districts = sorted(set(geo_districts + db_districts))

    # Merge DB villages with geo villages
    db_villages = []
    if district:
        db_villages = list(qs.filter(district_name=district)
                           .values_list('village', flat=True).distinct())
    all_villages = sorted(set(geo_villages + db_villages))

    return JsonResponse({
        'districts':    all_districts,
        'villages':     all_villages,
        'descriptions': descs,
    })

# ── GOVS SUMMARY ──────────────────────────────────────────────────
@login_required(login_url='login')
def govs_summary_api(request):
    stats = (Violation.objects.filter(status='approved')
             .values('governorate__pcode','governorate__name_ar','governorate__name_en')
             .annotate(count=Count('id'),total_area=Sum('area_total')).order_by('-count'))
    return JsonResponse({'governorates': list(stats)})

# ── ADD VIOLATION ─────────────────────────────────────────────────
@login_required(login_url='login')
def add_violation_api(request):
    role = get_role(request.user)
    if role not in ('data_entry','supervisor','manager'):
        return JsonResponse({'error':'غير مصرح لك بإضافة بيانات'},status=403)
    if request.method != 'POST':
        return JsonResponse({'error':'POST only'},status=405)
    try: data = json.loads(request.body)
    except: return JsonResponse({'error':'بيانات غير صالحة'},status=400)

    required = ['gov_pcode','district_name','village','occupant','basin','description','latitude','longitude']
    missing = [f for f in required if not str(data.get(f,'')).strip()]
    if missing:
        return JsonResponse({'error':f'حقول مطلوبة: {", ".join(missing)}'},status=400)

    gov_pcode = data.get('gov_pcode', '').strip()
    if not gov_pcode:
        return JsonResponse({'error': 'كود المحافظة مطلوب'}, status=400)

    # استرجاع المحافظة أو إنشاؤها تلقائياً من ملف الجيو
    gov = Governorate.objects.filter(pcode=gov_pcode).first()
    if not gov:
        _load_geo()
        gov_info   = next((g for g in (_GOVS or []) if g.get('pcode') == gov_pcode), None)
        gov_name_ar = gov_info['name_ar'] if gov_info else gov_pcode
        gov_name_en = gov_info.get('name_en', gov_pcode) if gov_info else gov_pcode
        gov = Governorate.objects.create(
            pcode=gov_pcode, name_ar=gov_name_ar,
            name_en=gov_name_en, has_data=True,
        )
    elif not gov.has_data:
        gov.has_data = True
        gov.save(update_fields=['has_data'])

    num  = Violation.objects.filter(governorate=gov).count() + 1
    code = f"{gov_pcode}-{num:04d}"
    status = 'approved' if role in ('supervisor','manager') else 'pending'

    v = Violation.objects.create(
        governorate=gov,
        district_name  = data['district_name'].strip(),
        village        = data['village'].strip(),
        village_pcode  = data.get('village_pcode',''),
        code           = code,
        occupant       = data['occupant'].strip(),
        basin          = data['basin'].strip(),
        description    = data['description'].strip(),
        area_outside   = float(data.get('area_outside',0) or 0),
        area_public    = float(data.get('area_public',0) or 0),
        area_nile_bank = float(data.get('area_nile_bank',0) or 0),
        area_total     = float(data.get('area_total',0) or 0),
        latitude       = float(data['latitude']),
        longitude      = float(data['longitude']),
        geo_exact      = bool(data.get('geo_exact',False)),
        field_notes    = data.get('field_notes',''),
        status         = status,
        submitted_by   = request.user,
        import_batch   = 'manual',
    )
    msg = 'تم الحفظ وإرسال للمراجعة' if status=='pending' else 'تم الحفظ والاعتماد'
    log_action(request, 'add', f'تعدٍّ: {v.code} — {v.village}')
    return JsonResponse({'success':True,'id':v.id,'code':v.code,'message':msg})

# ── EDIT VIOLATION ────────────────────────────────────────────────
@login_required(login_url='login')
def edit_violation_api(request, pk):
    role = get_role(request.user)
    if role not in ('data_entry','supervisor','manager'):
        return JsonResponse({'error':'غير مصرح'},status=403)
    v = get_object_or_404(Violation, pk=pk)
    # data_entry يمكنه تعديل سجلاته فقط (pending أو approved)
    # supervisor/manager يمكنهم تعديل أي سجل
    if role == 'data_entry':
        if v.submitted_by != request.user:
            return JsonResponse({'error': 'لا يمكنك تعديل سجل أدخله شخص آخر'}, status=403)
        if v.status == 'rejected':
            return JsonResponse({'error': 'لا يمكن تعديل سجل مرفوض — تواصل مع المشرف'}, status=403)
    try: data = json.loads(request.body)
    except: return JsonResponse({'error':'بيانات غير صالحة'},status=400)

    for f in ['district_name','village','occupant','basin','description','field_notes']:
        if f in data: setattr(v, f, str(data[f]).strip())
    for f in ['area_outside','area_public','area_nile_bank','area_total','latitude','longitude']:
        if f in data: setattr(v, f, float(data[f] or 0))
    if role=='data_entry': v.status='pending'
    v.save()
    return JsonResponse({'success':True,'message':'تم التعديل بنجاح'})

# ── APPROVE / REJECT ──────────────────────────────────────────────
@login_required(login_url='login')
def approve_violation_api(request, pk):
    role = get_role(request.user)
    if role not in ('supervisor','manager'):
        return JsonResponse({'error':'غير مصرح'},status=403)
    v = get_object_or_404(Violation, pk=pk)
    try: data = json.loads(request.body)
    except: data = {}
    v.reviewed_by  = request.user
    v.reviewed_at  = timezone.now()
    v.review_notes = data.get('notes','')
    v.status       = 'approved' if data.get('action','approve')=='approve' else 'rejected'
    v.save()
    log_action(request, 'approve' if v.status=='approved' else 'reject',
               f'{v.code} — {v.village}', v.review_notes)
    msg = 'تم اعتماد السجل' if v.status=='approved' else 'تم رفض السجل'
    return JsonResponse({'success':True,'message':msg,'status':v.status})

# ── NOTES ─────────────────────────────────────────────────────────
@login_required(login_url='login')
def notes_api(request, pk):
    v = get_object_or_404(Violation, pk=pk)
    if request.method=='GET':
        notes = list(v.notes.select_related('user').values(
            'id','text','user__username','user__first_name','created_at'))
        return JsonResponse({'notes':notes})
    if request.method=='POST':
        try: data = json.loads(request.body)
        except: return JsonResponse({'error':'بيانات غير صالحة'},status=400)
        text = data.get('text','').strip()
        if not text: return JsonResponse({'error':'الملاحظة فارغة'},status=400)
        note = ViolationNote.objects.create(violation=v,user=request.user,text=text)
        return JsonResponse({'success':True,'id':note.id,'text':note.text,
            'user':request.user.get_full_name() or request.user.username,
            'created_at':note.created_at.strftime('%Y-%m-%d %H:%M')})
    return JsonResponse({'error':'Method not allowed'},status=405)

# ── IMAGE UPLOAD ──────────────────────────────────────────────────
@login_required(login_url='login')
def upload_image_api(request, pk):
    role = get_role(request.user)
    if role not in ('data_entry','supervisor','manager'):
        return JsonResponse({'error':'غير مصرح'},status=403)
    if request.method!='POST':
        return JsonResponse({'error':'POST only'},status=405)
    v = get_object_or_404(Violation, pk=pk)
    images = request.FILES.getlist('images')
    if not images: return JsonResponse({'error':'لم يتم رفع أي صورة'},status=400)
    saved = []
    for img in images:
        vi = ViolationImage.objects.create(
            violation=v, image=img,
            caption=request.POST.get('caption',''),
            uploaded_by=request.user)
        saved.append({'id':vi.id,'url':vi.image.url})
    return JsonResponse({'success':True,'images':saved})

# ── VIOLATION DETAIL ──────────────────────────────────────────────
@login_required(login_url='login')
def violation_detail_api(request, pk):
    import traceback as tb
    role = get_role(request.user)

    try:
        v = get_object_or_404(Violation, pk=pk)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

    # safe getter
    def sg(obj, attr, default=''):
        try:    return getattr(obj, attr) or default
        except: return default

    def sf(obj, attr, default=0.0):
        try:    return float(getattr(obj, attr) or 0)
        except: return default

    try:
        status_val = sg(v, 'status', 'approved')
        if status_val == 'pending' and role not in ('supervisor', 'manager'):
            try:
                if v.submitted_by != request.user:
                    return JsonResponse({'error': 'غير مصرح'}, status=403)
            except Exception:
                pass

        try:
            images = [
                {'id': i.id, 'url': i.image.url, 'caption': sg(i, 'caption')}
                for i in v.images.all()
            ]
        except Exception:
            images = []

        try:
            sub_by = v.submitted_by.get_full_name() if v.submitted_by else ''
        except Exception:
            sub_by = ''

        try:
            sub_at = v.submitted_at.strftime('%Y-%m-%d %H:%M') if v.submitted_at else ''
        except Exception:
            sub_at = ''

        try:
            gov_pcode = v.governorate.pcode if v.governorate else ''
            gov_name  = v.governorate.name_ar if v.governorate else ''
        except Exception:
            gov_pcode = gov_name = ''

        data = {
            'id':           v.id,
            'code':         sg(v, 'code'),
            'status':       status_val,
            'gov_pcode':    gov_pcode,
            'gov_name':     gov_name,
            'district_name':sg(v, 'district_name'),
            'village':      sg(v, 'village'),
            'village_pcode':sg(v, 'village_pcode'),
            'occupant':     sg(v, 'occupant'),
            'basin':        sg(v, 'basin'),
            'description':  sg(v, 'description'),
            'area_outside': sf(v, 'area_outside'),
            'area_public':  sf(v, 'area_public'),
            'area_nile_bank': sf(v, 'area_nile_bank'),
            'area_total':   sf(v, 'area_total'),
            'latitude':     sf(v, 'latitude'),
            'longitude':    sf(v, 'longitude'),
            'geo_exact':    bool(sg(v, 'geo_exact', False)),
            'field_notes':  sg(v, 'field_notes'),
            'submitted_by': sub_by,
            'submitted_at': sub_at,
            'review_notes': sg(v, 'review_notes'),
            'images':       images,
            'notes': [],
            'role':         role,
        }
        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'trace': tb.format_exc()
        }, status=500)

# ── PENDING LIST ──────────────────────────────────────────────────
@login_required(login_url='login')
def pending_api(request):
    role = get_role(request.user)
    if role not in ('supervisor','manager'):
        return JsonResponse({'error':'غير مصرح'},status=403)
    qs = Violation.objects.filter(status='pending').select_related('governorate','submitted_by')
    records = list(qs.values('id','code','governorate__name_ar','district_name','village',
        'occupant','description','area_total','submitted_by__username',
        'submitted_by__first_name','submitted_at'))
    return JsonResponse({'records':records,'count':len(records)})


# ══════════════════════════════════════════════════════════════════
# AUDIT LOG HELPER
# ══════════════════════════════════════════════════════════════════
def log_action(request, action, target='', details=''):
    from .models import AuditLog
    ip = (request.META.get('HTTP_X_FORWARDED_FOR','').split(',')[0].strip()
          or request.META.get('REMOTE_ADDR',''))
    AuditLog.objects.create(
        user=request.user if request.user.is_authenticated else None,
        action=action, target=target, details=details, ip_address=ip or None,
    )


# ══════════════════════════════════════════════════════════════════
# ADMIN DASHBOARD
# ══════════════════════════════════════════════════════════════════
@login_required(login_url='login')
def admin_dashboard(request):
    role = get_role(request.user)
    if role != 'manager':
        return redirect('map')
    log_action(request, 'login', 'لوحة الإدارة')
    return render(request, 'violations/admin_dashboard.html', {'user': request.user})


@login_required(login_url='login')
def admin_stats_api(request):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    from .models import AuditLog
    from django.utils import timezone
    today = timezone.now().date()
    stats = {
        'total':      Violation.objects.count(),
        'approved':   Violation.objects.filter(status='approved').count(),
        'pending':    Violation.objects.filter(status='pending').count(),
        'rejected':   Violation.objects.filter(status='rejected').count(),
        'total_area': Violation.objects.aggregate(s=Sum('area_total'))['s'] or 0,
        'geo_count':  Violation.objects.filter(geo_exact=True).count(),
        'user_count': User.objects.filter(is_active=True).count(),
        'gov_count':  Governorate.objects.filter(has_data=True).count(),
        'today_logs': AuditLog.objects.filter(timestamp__date=today).count(),
        'top_govs': list(
            Violation.objects.filter(status='approved')
            .values('governorate__name_ar')
            .annotate(count=Count('id'))
            .order_by('-count')[:5]
            .values('count', name=models.F('governorate__name_ar'))
        ),
    }
    return JsonResponse(stats)


@login_required(login_url='login')
def admin_users_api(request, user_id=None):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)

    if request.method == 'GET':
        if user_id:
            u = get_object_or_404(User, pk=user_id)
            try:    role = u.profile.role
            except: role = 'viewer'
            try:    gov_pcode = u.profile.governorate.pcode if u.profile.governorate else ''
            except: gov_pcode = ''
            return JsonResponse({
                'id': u.id, 'username': u.username,
                'first_name': u.first_name, 'last_name': u.last_name,
                'email': u.email, 'role': role, 'gov_pcode': gov_pcode,
            })
        users = User.objects.select_related('profile__governorate').all().order_by('username')
        data = []
        for u in users:
            try:    role = u.profile.role
            except: role = 'viewer'
            try:    gov = u.profile.governorate.name_ar if u.profile.governorate else ''
            except: gov = ''
            data.append({
                'id': u.id, 'username': u.username,
                'full_name': u.get_full_name(), 'email': u.email,
                'role': role, 'governorate': gov,
                'is_active': u.is_active,
                'last_login': u.last_login.isoformat() if u.last_login else None,
            })
        return JsonResponse({'users': data})

    if request.method == 'POST':
        try: data = json.loads(request.body)
        except: return JsonResponse({'error': 'بيانات غير صالحة'}, status=400)

        if user_id:
            u = get_object_or_404(User, pk=user_id)
            u.first_name = data.get('first_name', u.first_name)
            u.last_name  = data.get('last_name',  u.last_name)
            u.email      = data.get('email',       u.email)
            if data.get('password'): u.set_password(data['password'])
            u.save()
        else:
            if not data.get('username'):
                return JsonResponse({'error': 'اسم المستخدم مطلوب'}, status=400)
            u, created = User.objects.get_or_create(username=data['username'])
            u.first_name = data.get('first_name', '')
            u.last_name  = data.get('last_name', '')
            u.email      = data.get('email', '')
            if data.get('password'): u.set_password(data['password'])
            elif created: u.set_password('Change@123')
            u.save()

        profile, _ = UserProfile.objects.get_or_create(user=u)
        profile.role = data.get('role', 'viewer')
        gov_pcode = data.get('gov_pcode', '')
        if gov_pcode:
            try: profile.governorate = Governorate.objects.get(pcode=gov_pcode)
            except: profile.governorate = None
        else:
            profile.governorate = None
        profile.save()
        log_action(request, 'edit', f'مستخدم: {u.username}')
        return JsonResponse({'success': True, 'message': f'تم حفظ المستخدم {u.username}'})


@login_required(login_url='login')
def admin_toggle_user_api(request, user_id):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    u = get_object_or_404(User, pk=user_id)
    if u == request.user:
        return JsonResponse({'error': 'لا يمكنك إيقاف حسابك'}, status=400)
    u.is_active = not u.is_active
    u.save()
    msg = f'تم {"تفعيل" if u.is_active else "إيقاف"} المستخدم {u.username}'
    log_action(request, 'edit', msg)
    return JsonResponse({'success': True, 'message': msg})


@login_required(login_url='login')
def admin_logs_api(request):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    from .models import AuditLog
    qs = AuditLog.objects.select_related('user').all()
    action = request.GET.get('action', '')
    user   = request.GET.get('user', '')
    date   = request.GET.get('date', '')
    limit  = int(request.GET.get('limit', 100))
    if action: qs = qs.filter(action=action)
    if user:   qs = qs.filter(user__username=user)
    if date:
        try:
            from datetime import date as dt
            d = dt.fromisoformat(date)
            qs = qs.filter(timestamp__date=d)
        except: pass
    logs = list(qs[:limit].values(
        'id','action','target','details','ip_address','timestamp',
        'user__username','user__first_name',
    ))
    users = list(AuditLog.objects.values_list('user__username', flat=True)
                 .distinct().exclude(user__username=None))
    return JsonResponse({
        'logs': [{
            'action': l['action'],
            'action_display': dict(AuditLog.ACTION_CHOICES).get(l['action'], l['action']),
            'target': l['target'],
            'details': l['details'],
            'ip': l['ip_address'],
            'user': l['user__first_name'] or l['user__username'] or 'غير معروف',
            'time': l['timestamp'].strftime('%Y-%m-%d %H:%M') if l['timestamp'] else '',
        } for l in logs],
        'users': users,
    })


@login_required(login_url='login')
def admin_logs_export(request):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    from .models import AuditLog
    import openpyxl
    from django.http import HttpResponse
    import io
    qs = AuditLog.objects.select_related('user').all()[:5000]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'سجل الأنشطة'
    ws.sheet_view.rightToLeft = True
    headers = ['التوقيت','المستخدم','الحدث','الهدف','التفاصيل','عنوان IP']
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    action_labels = dict(AuditLog.ACTION_CHOICES)
    for row_i, log in enumerate(qs, 2):
        ws.cell(row=row_i, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M') if log.timestamp else '')
        ws.cell(row=row_i, column=2, value=log.user.username if log.user else 'غير معروف')
        ws.cell(row=row_i, column=3, value=action_labels.get(log.action, log.action))
        ws.cell(row=row_i, column=4, value=log.target)
        ws.cell(row=row_i, column=5, value=log.details)
        ws.cell(row=row_i, column=6, value=log.ip_address or '')
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    response = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="audit_log.xlsx"'
    log_action(request, 'export', 'سجل الأنشطة')
    return response


@login_required(login_url='login')
def admin_govs_api(request):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    _load_geo()
    govs = sorted(_GOVS, key=lambda g: g['name_ar'])
    return JsonResponse({'govs': govs})


# ══════════════════════════════════════════════════════════════════
# EXPORT — Excel & PDF (for all users)
# ══════════════════════════════════════════════════════════════════
@login_required(login_url='login')
def export_excel_view(request):
    from .exports import export_excel
    from django.http import HttpResponse
    role = get_role(request.user)
    qs = Violation.objects.select_related('governorate').all()
    if role not in ('supervisor','manager'):
        qs = qs.filter(status='approved')

    gov      = request.GET.get('gov','')
    district = request.GET.get('district','')
    village  = request.GET.get('village','')
    status   = request.GET.get('status','')
    min_area = request.GET.get('min_area','')
    search   = request.GET.get('search','')

    filters = {}
    if gov:      qs = qs.filter(governorate__pcode=gov);  filters['gov'] = gov
    if district: qs = qs.filter(district_name=district);  filters['district'] = district
    if village:  qs = qs.filter(village=village);          filters['village'] = village
    if status and role in ('supervisor','manager'):
        qs = qs.filter(status=status); filters['status'] = status
    if min_area:
        try: qs = qs.filter(area_total__gte=float(min_area)); filters['min_area'] = min_area
        except: pass
    if search:
        qs = qs.filter(Q(occupant__icontains=search)|Q(code__icontains=search)|
                       Q(village__icontains=search))
        filters['search'] = search

    records = list(qs.values(
        'code','governorate__name_ar','district_name','village','occupant',
        'basin','description','area_nile_bank','area_public','area_outside',
        'area_total','status',
    ))

    buf = export_excel(records, filters, request.user)
    log_action(request, 'export', f'Excel — {len(records)} سجل')
    response = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="violations_{len(records)}.xlsx"'
    return response


@login_required(login_url='login')
def export_pdf_view(request):
    from .exports import export_pdf
    from django.http import HttpResponse
    role = get_role(request.user)
    qs = Violation.objects.select_related('governorate').all()
    if role not in ('supervisor','manager'):
        qs = qs.filter(status='approved')

    gov      = request.GET.get('gov','')
    district = request.GET.get('district','')
    status   = request.GET.get('status','')
    min_area = request.GET.get('min_area','')
    filters  = {}
    if gov:      qs = qs.filter(governorate__pcode=gov);  filters['gov'] = gov
    if district: qs = qs.filter(district_name=district);  filters['district'] = district
    if status and role in ('supervisor','manager'):
        qs = qs.filter(status=status); filters['status'] = status
    if min_area:
        try: qs = qs.filter(area_total__gte=float(min_area)); filters['min_area'] = min_area
        except: pass

    records = list(qs.values(
        'code','governorate__name_ar','district_name','village','occupant',
        'basin','description','area_nile_bank','area_public','area_outside',
        'area_total','status',
    ))

    buf = export_pdf(records, filters, request.user)
    log_action(request, 'export', f'PDF — {len(records)} سجل')
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="violations_{len(records)}.pdf"'
    return response

# ── REJECTED VIOLATIONS API ───────────────────────────────────────
@login_required(login_url='login')
def rejected_api(request):
    """قائمة التعديات المرفوضة — للمشرف والمدير فقط"""
    role = get_role(request.user)
    if role not in ('supervisor', 'manager'):
        return JsonResponse({'error': 'غير مصرح'}, status=403)

    qs = (Violation.objects
          .filter(status='rejected')
          .select_related('governorate', 'submitted_by', 'reviewed_by'))

    gov = request.GET.get('gov', '')
    if gov:
        qs = qs.filter(governorate__pcode=gov)

    records = list(qs.values(
        'id', 'code', 'governorate__name_ar', 'district_name', 'village',
        'occupant', 'description', 'area_total',
        'submitted_by__username', 'submitted_by__first_name',
        'reviewed_by__username', 'reviewed_by__first_name',
        'review_notes', 'submitted_at', 'reviewed_at',
    ))
    return JsonResponse({'records': records, 'count': len(records)})


# ── RESTORE REJECTED — إعادة سجل مرفوض للانتظار ─────────────────
@login_required(login_url='login')
def restore_violation_api(request, pk):
    """إعادة تعدٍّ مرفوض إلى حالة الانتظار"""
    role = get_role(request.user)
    if role not in ('supervisor', 'manager'):
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    v = get_object_or_404(Violation, pk=pk, status='rejected')
    v.status       = 'pending'
    v.review_notes = ''
    v.reviewed_by  = None
    v.reviewed_at  = None
    v.save()
    log_action(request, 'edit', f'إعادة مرفوض للانتظار: {v.code}')
    return JsonResponse({'success': True, 'message': f'تمت إعادة {v.code} إلى الانتظار'})
